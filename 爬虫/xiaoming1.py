import openpyxl

# 加载原始Excel文件
excel = openpyxl.load_workbook('demo.xlsx')
sheet = excel['Sheet']
rows = sheet.max_row

# 创建新的Excel文件和工作表
new_excel = openpyxl.Workbook()
new_sheet = new_excel.active
new_sheet.title = 'High Scores'
new_sheet.append(['班级', '姓名', '数学成绩'])  # 添加标题行

# 遍历原始数据
for i in range(2, rows + 1):
    classroom = sheet.cell(i, 1).value
    name = sheet.cell(i, 2).value
    point = sheet.cell(i, 3).value

    try:
        if int(point) >= 90:
            # 将符合条件的记录添加到新表
            new_sheet.append([classroom, name, point])
    except ValueError:
        print(f"警告：第 {i} 行数学成绩「{point}」不是有效数字，已跳过")

# 保存新的Excel文件
new_excel.save('score.xlsx')

# 关闭文件（非必须，save()会自动保存）
excel.close()
new_excel.close()

print("处理完成！高分学生已保存到 score.xlsx")
        







