from openpyxl import Workbook,load_workbook
from openpyxl import styles
import os

path=os.path.join(os.path.dirname(os.path.abspath(__file__)),'统计.xlsx')
wb=load_workbook(path)
ws=wb['kakaka']
for row in ws:
    for cell in row:
        print(cell.value,end=',')
    print()

my_font=styles.Font(name='等线',size=24,italic=True,color='FF0000',bold=True)
ws['A2'].font=my_font
wb.save(path)