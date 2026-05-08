from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText
from email.header import Header
import os
from email.utils import formataddr
import time

path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '2.xlsx')
wb = load_workbook(path)
ws = wb['Sheet1']

# 提取表头
table_col_html = "<thead><tr>"
for cell in ws[1]:  # 标题行固定为第1行
    table_col_html += f"<th>{cell.value}</th>"
table_col_html += "</tr></thead>"

# 遍历数据行（从第2行开始，跳过标题）
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    # 跳过空行或无效邮箱
    if not row[0].value or not row[3].value:
        print(f"跳过无效行：{row}")
        continue
    
    # 构建数据行HTML
    row_text = "<tr>"
    for cell in row:
        row_text += f"<td>{cell.value}</td>"
    row_text += "</tr>"
    
    name = row[0].value
    addr = row[3].value
    
    # 构建完整邮件内容
    mail_body_context = f"""
    <html>
    <body>
        <table border="1px solid black">
            {table_col_html}
            <tbody>
                {row_text}
            </tbody>
        </table>
    </body>
    </html>
    """
    
    # 创建并发送邮件
    msg = MIMEText(mail_body_context, 'html', 'utf-8')
    msg['From'] = formataddr(('二年级年级组', '1281545400@qq.com'))
    msg['To'] = formataddr((name, addr))
    msg['Subject'] = Header('二年级成绩', 'utf-8')
    
    try:
        with smtplib.SMTP_SSL('smtp.qq.com', 465) as smtp_obj:
            smtp_obj.login('1281545400@qq.com', 'qzgiebdeoprmhfgc')
            smtp_obj.sendmail('1281545400@qq.com', [addr], msg.as_string())
            print(f'发送成功至 {name} ({addr})')
            time.sleep(5)
    except smtplib.SMTPException as e:
        print(f'发送失败至 {addr}，SMTP错误：{e}')
    except Exception as e:
        print(f'发送失败至 {addr}，未知错误：{e}')